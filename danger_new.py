import argparse
from datetime import datetime, timedelta
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


parser = argparse.ArgumentParser()
parser.add_argument("--year", type=int, default=2026)
parser.add_argument("--month", type=int, default=5)
parser.add_argument("--input", type=str, default=None, help="入力Excelパス。未指定なら shift/YYYY_MM_shift.xlsx")
parser.add_argument("--output", type=str, default=None, help="出力Excelパス。未指定なら shift/YYYY_MM_shift_danger.xlsx")
parser.add_argument("--keep_monthly", action="store_true", help="月集計シートも残す場合に指定")
args = parser.parse_args()


year = args.year
month = args.month
month_str = f"{year}_{month:02d}"

file_path = Path(args.input) if args.input else Path(f"shift/{month_str}_shift.xlsx")
output_path = Path(args.output) if args.output else Path(f"shift/{month_str}_shift_danger.xlsx")


# =========================
# 30分スロット生成
# =========================
def generate_slots(start="06:00", end="20:30"):
    slots = []
    current = datetime.strptime(start, "%H:%M")
    end_dt = datetime.strptime(end, "%H:%M")

    while current < end_dt:
        nxt = current + timedelta(minutes=30)
        slots.append(current.strftime("%H:%M") + "-" + nxt.strftime("%H:%M"))
        current = nxt

    return slots


all_slot_names = generate_slots()


# =========================
# 日別シート判定
# 例：2026-05-01 のようなシートだけ対象
# =========================
def is_daily_sheet(sheet_name: str) -> bool:
    try:
        datetime.strptime(sheet_name, "%Y-%m-%d")
        return True
    except ValueError:
        return False


# =========================
# 日別シートから勤務部分だけ取得
# =========================
def get_slot_df(excel_path: Path, sheet_name: str) -> pd.DataFrame:
    df = pd.read_excel(excel_path, sheet_name=sheet_name, header=None)

    # 0行目: 日付・曜日
    # 1行目: 氏名・時間
    # 2行目以降: 従業員行 + 下部集計行
    employee_df = df.iloc[2:, :].copy()

    # 下部集計行を除外
    summary_labels = [
        "社員S時間",
        "PT時間",
        "合計時間",
        "追加入時間",
        "総人数",
    ]

    employee_df = employee_df[~employee_df.iloc[:, 0].isin(summary_labels)]

    # 空行を除外
    employee_df = employee_df[employee_df.iloc[:, 0].notna()]

    # スロット列だけ取得
    slot_df = employee_df.iloc[:, 1:1 + len(all_slot_names)]

    return slot_df


# =========================
# 入力チェック
# =========================
if not file_path.exists():
    raise FileNotFoundError(f"入力ファイルが見つかりません: {file_path}")

xls = pd.ExcelFile(file_path)
sheet_names = xls.sheet_names


# =========================
# 危険・注意スロット抽出
# =========================
danger_slots = []   # 1人以下
warning_slots = []  # ちょうど2人

daily_sheet_names = [sheet for sheet in sheet_names if is_daily_sheet(sheet)]

for sheet in daily_sheet_names:
    slot_df = get_slot_df(file_path, sheet)

    for i, slot_name in enumerate(all_slot_names):
        if i >= slot_df.shape[1]:
            continue

        count = (slot_df.iloc[:, i] == "□").sum()

        if count <= 1:
            danger_slots.append({
                "date": sheet,
                "slot": slot_name,
                "staff_count": int(count),
            })
        elif count == 2:
            warning_slots.append({
                "date": sheet,
                "slot": slot_name,
                "staff_count": int(count),
            })


# =========================
# コンソール出力
# =========================
warning_df = pd.DataFrame(warning_slots)
print("🟡【注意】2人のスロット")
if warning_df.empty:
    print("なし（問題ありません）")
else:
    warning_df = warning_df.sort_values(["date", "slot"])
    print(warning_df.to_string(index=False))


danger_df = pd.DataFrame(danger_slots)
print("\n🔴【危険】1人以下のスロット")
if danger_df.empty:
    print("なし（問題ありません）")
else:
    danger_df = danger_df.sort_values(["date", "slot"])
    print(danger_df.to_string(index=False))


# =========================
# 危険日だけ残す
# =========================
danger_dates = set(danger_df["date"].tolist()) if not danger_df.empty else set()

# 元Excelをそのまま開くので、書式・列幅・行高・結合セルを維持できる
wb = load_workbook(file_path)

keep_sheets = set(danger_dates)

# 月集計シートも残したい場合のみ残す
if args.keep_monthly:
    for monthly_sheet in ["月", "月間シフト表"]:
        if monthly_sheet in wb.sheetnames:
            keep_sheets.add(monthly_sheet)

# 危険日以外のシートを削除
for sheet in wb.sheetnames[:]:
    if sheet not in keep_sheets:
        del wb[sheet]

# 危険日が1つもない場合、Excelは最低1シート必要なので案内シートを作る
if not wb.sheetnames:
    ws = wb.create_sheet("危険日なし")
    ws["A1"] = "1人以下の危険スロットがある日はありません。"

output_path.parent.mkdir(parents=True, exist_ok=True)
wb.save(output_path)

print(f"\n不足日のみ出力完了：{output_path}")
